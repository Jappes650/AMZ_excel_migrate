import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import json
from openpyxl import load_workbook
import re

class ExcelMigrationTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Migrationstool - Verbesserte Version")
        self.root.geometry("1000x800")
        self.root.configure(bg='#f0f0f0')
        
        # Style konfigurieren
        style = ttk.Style()
        style.theme_use('clam')
        
        # Variablen
        self.template_file = tk.StringVar()
        self.data_file = tk.StringVar()
        self.template_sheet = tk.StringVar()
        self.data_sheet = tk.StringVar()
        self.mapping = {}
        self.available_data_columns = set()
        
        # Standardwerte für Zeilenkonfiguration
        self.data_header_row = tk.IntVar(value=3)
        self.data_start_row = tk.IntVar(value=4)
        self.template_header_row = tk.IntVar(value=5)
        self.template_start_row = tk.IntVar(value=8)
        
        # Listen für verfügbare Sheets
        self.template_sheets = []
        self.data_sheets = []
        
        # Erweiterte Keyword-basierte Zuordnungen
        self.keyword_mappings = {
            # Basis-Mappings
            'list_price_with_tax': r'uvp_list_price\[marketplace_id=.*\]#1\.value',
            'list_price_with_tax': r'list_price\[marketplace_id=.*\]#1\.value_with_tax',
            'relationship_type': r'child_parent_sku_relationship\[marketplace_id=.*\]#1\.child_relationship_type',
            'parent_child': r'parentage_level\[marketplace_id=.*\]#1\.value',
            'parent_sku': r'child_parent_sku_relationship\[marketplace_id=.*\]#1\.parent_sku',
            'update_delete': '::record_action',
            'item_sku': r'contribution_sku#1\.value',
            'external_product_id_type': r'amzn1\.volt\.ca\.product_id_type',
            'external_product_id': r'amzn1\.volt\.ca\.product_id_value',
            'quantity': r'fulfillment_availability#1\.quantity',
            'standard_price': r'purchasable_offer\[marketplace_id=.*\]\[audience=ALL\]#1\.our_price#1\.schedule#1\.value_with_tax',
            'fulfillment_center_id': r'fulfillment_availability#1\.fulfillment_channel_code',
            
            # Neue/korrigierte Mappings
            'asin': r'amzn1\.volt\.ca\.product_id_value',  # ASIN mapping
            'efficiency': r'efficiency\[marketplace_id=.*\]#1\.value',  # Efficiency mapping
            'feed_product_type': r'product_type#\d+\.value',  # Product type
            'fulfillment_latency': r'fulfillment_latency\[marketplace_id=.*\]#1\.value',  # Fulfillment latency
            'is_fragile': r'fragile\[marketplace_id=.*\]#1\.value',  # Fragile mapping
            'variation_theme': r'variation_theme#\d+\.name',  # variation Theme
            'wattage': r'wattage\[marketplace_id=.*\]#1\.value',  # Wattage mapping
            
            # Dimensions mappings
            'depth_front_to_back': r'item_depth_width_height\[marketplace_id=.*\]#1\.depth\.value',
            'depth_front_to_back_unit_of_measure': r'item_depth_width_height\[marketplace_id=.*\]#1\.depth\.unit',
            'depth_width_side_to_side': r'item_depth_width_height\[marketplace_id=.*\]#1\.height\.value',
            'depth_width_side_to_side_unit_of_measure': r'item_depth_width_height\[marketplace_id=.*\]#1\.height\.unit',
            'depth_height_floor_to_top': r'item_depth_width_height\[marketplace_id=.*\]#1\.width\.value',
            'depth_height_floor_to_top_unit_of_measure': r'item_depth_width_height\[marketplace_id=.*\]#1\.width\.unit',
            
            # Package dimensions
            'package_length_unit_of_measure': r'item_package_dimensions\[marketplace_id=.*\]#1\.length\.unit',
            'package_width_unit_of_measure': r'item_package_dimensions\[marketplace_id=.*\]#1\.width\.unit', 
            'package_height_unit_of_measure': r'item_package_dimensions\[marketplace_id=.*\]#1\.height\.unit',
            'package_length': r'item_package_dimensions\[marketplace_id=.*\]#1\.length\.value',
            'package_width': r'item_package_dimensions\[marketplace_id=.*\]#1\.width\.value',
            'package_height': r'item_package_dimensions\[marketplace_id=.*\]#1\.height\.value',
            
            # Package weight
            'package_weight': r'item_package_weight\[marketplace_id=[^]]+\]#\d+\.value',
            'package_weight_unit_of_measure': r'item_package_weight\[marketplace_id=[^]]+\]#\d+\.unit',

            
            # Item weight
            'item_weight': r'item_weight\[marketplace_id=.*\]#1\.value',
            'item_weight_unit_of_measure': r'item_weight\[marketplace_id=.*\]#1\.unit',
            
            # Product information
            'brand_name': r'brand\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'item_name': r'item_name\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'product_description': r'product_description\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'recommended_browse_nodes1': r'recommended_browse_nodes\[marketplace_id=.*\]#1\.value',
            'country_of_origin': r'country_of_origin\[marketplace_id=.*\]#1\.value',
            'condition_type': r'condition_type\[marketplace_id=.*\]#1\.value',
            'model_name': r'model_name\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'bullet_point1': r'bullet_point\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'bullet_point2': r'bullet_point\[marketplace_id=.*\]\[language_tag=.*\]#2\.value',
            'bullet_point3': r'bullet_point\[marketplace_id=.*\]\[language_tag=.*\]#3\.value',
            'bullet_point4': r'bullet_point\[marketplace_id=.*\]\[language_tag=.*\]#4\.value',
            'bullet_point5': r'bullet_point\[marketplace_id=.*\]\[language_tag=.*\]#5\.value',
            'size_name': r'size\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'color_name': r'color\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'generic_keywords': r'generic_keyword\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'merchant_shipping_group_name': r'merchant_shipping_group\[marketplace_id=.*\]#1\.value',
            'part_number': r'part_number\[marketplace_id=.*\]#1\.value',
            'manufacturer': r'manufacturer\[marketplace_id=.*\]\[language_tag=.*\]#1\.value',
            'model': r'model_number\[marketplace_id=.*\]#1\.value',
            'main_image_url': r'main_offer_image_locator\[marketplace_id=.*\]#1\.media_location',
            'other_image_url1': r'other_offer_image_locator_1\[marketplace_id=.*\]#1\.media_location',
            'other_image_url2': r'other_offer_image_locator_2\[marketplace_id=.*\]#1\.media_location',
            'other_image_url3': r'other_offer_image_locator_3\[marketplace_id=.*\]#1\.media_location',
            'other_image_url4': r'other_offer_image_locator_4\[marketplace_id=.*\]#1\.media_location',
            'other_image_url5': r'other_offer_image_locator_5\[marketplace_id=.*\]#1\.media_location',
            'condition_note': r'condition_note\[marketplace_id=.*\]#1\.value',
            'item_package_quantity': r'item_package_quantity\[marketplace_id=.*\]#1\.value',
            'product_tax_code': r'product_tax_code\[marketplace_id=.*\]#1\.value',
            'map_price': r'uvp_list_price\[marketplace_id=.*\]#1\.value',
            'merchant_release_date': r'merchant_release_date\[marketplace_id=.*\]#1\.value',
            'number_of_items': r'number_of_items\[marketplace_id=.*\]#1\.value',
            'number_of_boxes': r'number_of_items\[marketplace_id=.*\]#1\.value',  # Alternative für number_of_boxes
            'fulfillment_latency': r'fulfillment_availability#\d+\.lead_time_to_ship_max_days',
            'mounting_type': r'mounting_type\[marketplace_id=.*?\]\[language_tag=.*?\]#1\.value',
            'power_plug_type': r'power_plug_type\[marketplace_id=.*?\]#\d+\.value',
            'heat_output': r'heat_output\[marketplace_id=.*?\]#\d+\.value',
            'heating_method': r'heating_method\[marketplace_id=.*?\]\[language_tag=.*?\]#\d+\.value',
            'accepted_voltage_frequency': r'accepted_voltage_frequency\[marketplace_id=.*?\]#\d+\.value'
        }
        
        self.setup_ui()
        
    def setup_ui(self):
        # Hauptframe mit Notebook für Tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tab 1: Dateien & Konfiguration
        self.setup_files_tab()
        
        # Tab 2: Mapping
        self.setup_mapping_tab()
        
        # Tab 3: Migration
        self.setup_migration_tab()
        
    def setup_files_tab(self):
        files_frame = ttk.Frame(self.notebook)
        self.notebook.add(files_frame, text="1. Dateien & Konfiguration")
        
        # Hauptcontainer mit Padding
        main_container = ttk.Frame(files_frame, padding="20")
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Dateien Section
        files_section = ttk.LabelFrame(main_container, text="Dateien auswählen", padding="15")
        files_section.pack(fill=tk.X, pady=(0, 20))
        
        # Template Datei
        ttk.Label(files_section, text="Vorlage Datei:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=5)
        template_frame = ttk.Frame(files_section)
        template_frame.grid(row=0, column=1, columnspan=2, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        template_frame.columnconfigure(0, weight=1)
        
        self.template_entry = ttk.Entry(template_frame, textvariable=self.template_file, width=60)
        self.template_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        ttk.Button(template_frame, text="Durchsuchen", command=self.browse_template).grid(row=0, column=1)
        
        ttk.Label(files_section, text="Vorlage Tabellenblatt:", font=('Arial', 10)).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.template_sheet_combo = ttk.Combobox(files_section, textvariable=self.template_sheet, width=47, state="readonly")
        self.template_sheet_combo.grid(row=1, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        self.template_sheet_combo.bind('<<ComboboxSelected>>', self.on_template_sheet_select)
        
        # Daten Datei
        ttk.Label(files_section, text="Daten Datei:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky=tk.W, pady=5)
        data_frame = ttk.Frame(files_section)
        data_frame.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        data_frame.columnconfigure(0, weight=1)
        
        self.data_entry = ttk.Entry(data_frame, textvariable=self.data_file, width=60)
        self.data_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        ttk.Button(data_frame, text="Durchsuchen", command=self.browse_data).grid(row=0, column=1)
        
        ttk.Label(files_section, text="Daten Tabellenblatt:", font=('Arial', 10)).grid(row=3, column=0, sticky=tk.W, pady=5)
        self.data_sheet_combo = ttk.Combobox(files_section, textvariable=self.data_sheet, width=47, state="readonly")
        self.data_sheet_combo.grid(row=3, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        self.data_sheet_combo.bind('<<ComboboxSelected>>', self.on_data_sheet_select)
        
        files_section.columnconfigure(1, weight=1)
        
        # Konfiguration Section
        config_section = ttk.LabelFrame(main_container, text="Zeilenkonfiguration", padding="15")
        config_section.pack(fill=tk.X, pady=(0, 20))
        
        # Daten Konfiguration
        data_config = ttk.LabelFrame(config_section, text="Daten-Datei", padding="10")
        data_config.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        ttk.Label(data_config, text="Header Zeile:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Spinbox(data_config, from_=1, to=100, textvariable=self.data_header_row, width=10).grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(data_config, text="Start Zeile:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Spinbox(data_config, from_=1, to=100, textvariable=self.data_start_row, width=10).grid(row=1, column=1, padx=5, pady=2)
        
        # Template Konfiguration
        template_config = ttk.LabelFrame(config_section, text="Vorlage-Datei", padding="10")
        template_config.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))
        
        ttk.Label(template_config, text="Header Zeile:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Spinbox(template_config, from_=1, to=100, textvariable=self.template_header_row, width=10).grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(template_config, text="Start Zeile:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Spinbox(template_config, from_=1, to=100, textvariable=self.template_start_row, width=10).grid(row=1, column=1, padx=5, pady=2)
        
        config_section.columnconfigure(0, weight=1)
        config_section.columnconfigure(1, weight=1)
        
        # Datenvalidierung Section
        validation_section = ttk.LabelFrame(main_container, text="Datenvalidierung", padding="15")
        validation_section.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Button(validation_section, text="Verfügbare Daten prüfen", 
                  command=self.validate_data, style='Accent.TButton').pack(side=tk.LEFT, padx=(0, 10))
        
        self.validation_label = ttk.Label(validation_section, text="Noch nicht geprüft", foreground='gray')
        self.validation_label.pack(side=tk.LEFT)
        
        # Navigation
        nav_frame = ttk.Frame(main_container)
        nav_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(nav_frame, text="Weiter zum Mapping →", 
                  command=lambda: self.notebook.select(1), 
                  style='Accent.TButton').pack(side=tk.RIGHT)
        
    def setup_mapping_tab(self):
        mapping_frame = ttk.Frame(self.notebook)
        self.notebook.add(mapping_frame, text="2. Mapping")
        
        main_container = ttk.Frame(mapping_frame, padding="20")
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Info Label
        info_label = ttk.Label(main_container, 
                              text="Ordnen Sie die Spalten aus Ihrer Daten-Datei den entsprechenden Spalten in der Vorlage zu.",
                              font=('Arial', 10))
        info_label.pack(anchor=tk.W, pady=(0, 20))
        
        # Mapping Controls
        controls_frame = ttk.Frame(main_container)
        controls_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(controls_frame, text="Automatisches Mapping", 
                  command=self.auto_mapping_standalone, 
                  style='Accent.TButton').pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(controls_frame, text="Mapping anzeigen/bearbeiten", 
                  command=self.show_mapping).pack(side=tk.LEFT, padx=(0, 10))
        
        # Mapping Status
        self.mapping_status = ttk.Label(main_container, text="Kein Mapping erstellt", foreground='gray')
        self.mapping_status.pack(anchor=tk.W, pady=(0, 20))
        
        # Navigation
        nav_frame = ttk.Frame(main_container)
        nav_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(nav_frame, text="← Zurück zu Dateien", 
                  command=lambda: self.notebook.select(0)).pack(side=tk.LEFT)
        ttk.Button(nav_frame, text="Weiter zur Migration →", 
                  command=lambda: self.notebook.select(2), 
                  style='Accent.TButton').pack(side=tk.RIGHT)
        
    def setup_migration_tab(self):
        migration_frame = ttk.Frame(self.notebook)
        self.notebook.add(migration_frame, text="3. Migration")
        
        main_container = ttk.Frame(migration_frame, padding="20")
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Info Label
        info_label = ttk.Label(main_container, 
                              text="Starten Sie die Migration Ihrer Daten in die Vorlage.",
                              font=('Arial', 10))
        info_label.pack(anchor=tk.W, pady=(0, 20))
        
        # Migration Section
        migration_section = ttk.LabelFrame(main_container, text="Migration starten", padding="15")
        migration_section.pack(fill=tk.X, pady=(0, 20))
        
        # Fortschrittsanzeige
        self.progress = ttk.Progressbar(migration_section, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=(0, 10))
        
        # Status
        self.status_label = ttk.Label(migration_section, text="Bereit für Migration", foreground='green')
        self.status_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Start Button
        self.migration_button = ttk.Button(migration_section, text="Migration starten", 
                                          command=self.start_migration, 
                                          style='Accent.TButton')
        self.migration_button.pack(pady=10)
        
        # Debug Section
        debug_section = ttk.LabelFrame(main_container, text="Diagnose & Debug", padding="15")
        debug_section.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Button(debug_section, text="Debug-Informationen anzeigen", 
                  command=self.show_debug_info).pack(side=tk.LEFT, padx=(0, 10))
        
        # Navigation
        nav_frame = ttk.Frame(main_container)
        nav_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(nav_frame, text="← Zurück zum Mapping", 
                  command=lambda: self.notebook.select(1)).pack(side=tk.LEFT)
        
    def browse_template(self):
        filename = filedialog.askopenfilename(
            title="Vorlage Datei auswählen",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")]
        )
        if filename:
            self.template_file.set(filename)
            self.load_sheet_names(filename, 'template')
            
    def browse_data(self):
        filename = filedialog.askopenfilename(
            title="Daten Datei auswählen",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")]
        )
        if filename:
            self.data_file.set(filename)
            self.load_sheet_names(filename, 'data')
            
    def load_sheet_names(self, filename, file_type):
        try:
            if file_type == 'template':
                xl = pd.ExcelFile(filename)
                self.template_sheets = xl.sheet_names
                self.template_sheet_combo['values'] = self.template_sheets
                if self.template_sheets:
                    self.template_sheet.set(self.template_sheets[0])
            else:
                xl = pd.ExcelFile(filename)
                self.data_sheets = xl.sheet_names
                self.data_sheet_combo['values'] = self.data_sheets
                if self.data_sheets:
                    self.data_sheet.set(self.data_sheets[0])
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Lesen der Tabellenblätter: {str(e)}")
            
    def on_template_sheet_select(self, event):
        pass
        
    def on_data_sheet_select(self, event):
        # Reset validation when sheet changes
        self.validation_label.config(text="Noch nicht geprüft", foreground='gray')
        self.available_data_columns.clear()
        
    def validate_data(self):
        """Prüft verfügbare Daten ab Startzeile + 3 Zeilen"""
        if not self.data_file.get() or not self.data_sheet.get():
            messagebox.showerror("Fehler", "Bitte wählen Sie zuerst die Daten-Datei und das Tabellenblatt aus.")
            return
            
        try:
            # Lade Daten ab der konfigurierten Startzeile + 3
            check_start_row = self.data_start_row.get() + 2  # +2 weil 0-basiert und +3 Zeilen
            df = pd.read_excel(
                self.data_file.get(),
                sheet_name=self.data_sheet.get(),
                header=self.data_header_row.get()-1,
                nrows=5  # Prüfe nur 5 Zeilen ab der Startzeile
            )
            
            # Überspringe Zeilen bis zur Prüfzeile
            if check_start_row < len(df):
                check_df = df.iloc[check_start_row:]
            else:
                check_df = df
            
            # Finde Spalten mit Daten
            columns_with_data = []
            for column in check_df.columns:
                if check_df[column].notna().any():  # Hat mindestens einen nicht-leeren Wert
                    columns_with_data.append(column)
                    
            self.available_data_columns = set(columns_with_data)
            
            # Update UI
            count = len(columns_with_data)
            if count > 0:
                self.validation_label.config(
                    text=f"✓ {count} Spalten mit Daten gefunden", 
                    foreground='green'
                )
            else:
                self.validation_label.config(
                    text="⚠ Keine Daten gefunden", 
                    foreground='orange'
                )
                
            # Debug-Ausgabe
            print(f"Verfügbare Spalten mit Daten: {sorted(columns_with_data)}")
            
        except Exception as e:
            self.validation_label.config(text=f"Fehler bei Validierung: {str(e)}", foreground='red')
            messagebox.showerror("Fehler", f"Fehler bei der Datenvalidierung: {str(e)}")
            
    def auto_mapping_standalone(self):
        """Erstellt automatisches Mapping ohne separates Fenster"""
        if not self.template_file.get() or not self.data_file.get():
            messagebox.showerror("Fehler", "Bitte wählen Sie beide Dateien aus.")
            return
            
        if not self.template_sheet.get() or not self.data_sheet.get():
            messagebox.showerror("Fehler", "Bitte wählen Sie Tabellenblätter aus.")
            return
            
        try:
            # Lade Header aus beiden Dateien
            template_headers = self.read_headers(
                self.template_file.get(), 
                self.template_header_row.get(),
                self.template_sheet.get()
            )
            data_headers = self.read_headers(
                self.data_file.get(), 
                self.data_header_row.get(),
                self.data_sheet.get()
            )
            
            # Erstelle automatisches Mapping
            self.mapping = {}
            
            # Zuerst exakte Übereinstimmungen
            for data_header in data_headers:
                if data_header in template_headers:
                    # Nur hinzufügen wenn Daten verfügbar sind (falls validiert)
                    if not self.available_data_columns or data_header in self.available_data_columns:
                        self.mapping[data_header] = data_header
            
            # Dann Keyword-basiertes Mapping
            for data_header in data_headers:
                if data_header in self.mapping:
                    continue
                    
                # Nur verarbeiten wenn Daten verfügbar sind
                if self.available_data_columns and data_header not in self.available_data_columns:
                    continue
                    
                # Prüfe auf Keyword-Mapping
                for keyword, template_pattern in self.keyword_mappings.items():
                    if keyword.lower() in data_header.lower():
                        best_match = self.find_template_match(template_pattern, template_headers)
                        if best_match and best_match not in self.mapping.values():
                            self.mapping[data_header] = best_match
                            break
            
            # Update Status
            count = len(self.mapping)
            if count > 0:
                self.mapping_status.config(
                    text=f"✓ Automatisches Mapping erstellt: {count} Zuordnungen", 
                    foreground='green'
                )
            else:
                self.mapping_status.config(
                    text="⚠ Kein automatisches Mapping möglich", 
                    foreground='orange'
                )
                
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim automatischen Mapping: {str(e)}")
            
    def find_template_match(self, pattern, template_headers):
        """Findet die beste Übereinstimmung für ein Template-Pattern"""
        best_match = None
        best_score = 0
        
        for template_header in template_headers:
            score = self.pattern_match_score(pattern, template_header)
            if score > best_score:
                best_score = score
                best_match = template_header
                
        return best_match if best_score > 0 else None
        
    def show_mapping(self):
        if not self.template_file.get() or not self.data_file.get():
            messagebox.showerror("Fehler", "Bitte wählen Sie beide Dateien aus.")
            return
            
        if not self.template_sheet.get() or not self.data_sheet.get():
            messagebox.showerror("Fehler", "Bitte wählen Sie Tabellenblätter aus.")
            return
            
        try:
            # Lade Header aus beiden Dateien
            template_headers = self.read_headers(
                self.template_file.get(), 
                self.template_header_row.get(),
                self.template_sheet.get()
            )
            data_headers = self.read_headers(
                self.data_file.get(), 
                self.data_header_row.get(),
                self.data_sheet.get()
            )
            
            # Erstelle Mapping-Fenster
            mapping_window = tk.Toplevel(self.root)
            mapping_window.title("Spalten-Mapping bearbeiten")
            mapping_window.geometry("1200x800")
            mapping_window.grab_set()  # Modal
            
            # Erstelle Frame für das Mapping
            mapping_frame = ttk.Frame(mapping_window, padding="20")
            mapping_frame.pack(fill=tk.BOTH, expand=True)
            
            # Info Label
            info_frame = ttk.Frame(mapping_frame)
            info_frame.pack(fill=tk.X, pady=(0, 15))
            
            ttk.Label(info_frame, text="Spalten-Zuordnung", font=('Arial', 12, 'bold')).pack(anchor=tk.W)
            ttk.Label(info_frame, text="Doppelklicken Sie auf eine Spalte, um sie auszuwählen. Wählen Sie eine Spalte aus jeder Liste und klicken Sie 'Zuordnung hinzufügen'").pack(anchor=tk.W)
            
            # Status für aktuelle Auswahl
            selection_frame = ttk.Frame(info_frame)
            selection_frame.pack(fill=tk.X, pady=(10, 0))
            
            ttk.Label(selection_frame, text="Aktuelle Auswahl:", font=('Arial', 10, 'bold')).pack(anchor=tk.W)
            self.current_data_selection = ttk.Label(selection_frame, text="Daten-Spalte: Keine ausgewählt", foreground='gray')
            self.current_data_selection.pack(anchor=tk.W)
            self.current_template_selection = ttk.Label(selection_frame, text="Vorlage-Spalte: Keine ausgewählt", foreground='gray')
            self.current_template_selection.pack(anchor=tk.W)
            
            # Drei-Spalten Layout
            columns_frame = ttk.Frame(mapping_frame)
            columns_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
            
            # Daten-Spalten (Links)
            data_frame = ttk.LabelFrame(columns_frame, text="Daten-Spalten", padding="10")
            data_frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E), padx=(0, 10))
            
            data_listbox = tk.Listbox(data_frame, height=25, width=35, selectmode=tk.SINGLE)
            data_listbox.pack(fill=tk.BOTH, expand=True)
            
            # Scrollbar für Daten-Liste
            data_scrollbar = ttk.Scrollbar(data_frame, orient=tk.VERTICAL, command=data_listbox.yview)
            data_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            data_listbox.config(yscrollcommand=data_scrollbar.set)
            
            # Variablen für Auswahl
            self.selected_data_col = None
            self.selected_template_col = None
            
            # Fülle Daten-Header mit farblicher Kennzeichnung
            for i, header in enumerate(data_headers):
                data_listbox.insert(tk.END, header)
    
                if self.available_data_columns:
                    if header not in self.available_data_columns:
                        # Keine Daten verfügbar - grau kennzeichnen
                        data_listbox.itemconfig(i, {'fg': 'gray'})
                    elif header not in self.mapping:
                        # Hat Daten aber ist nicht gemappt - orange/rot kennzeichnen
                        data_listbox.itemconfig(i, {'fg': 'orange'})
                    
            # Template-Spalten (Mitte)
            template_frame = ttk.LabelFrame(columns_frame, text="Vorlage-Spalten", padding="10")
            template_frame.grid(row=0, column=1, sticky=(tk.N, tk.S, tk.W, tk.E), padx=(0, 10))
            
            template_listbox = tk.Listbox(template_frame, height=25, width=40, selectmode=tk.SINGLE)
            template_listbox.pack(fill=tk.BOTH, expand=True)
            
            # Scrollbar für Template-Liste
            template_scrollbar = ttk.Scrollbar(template_frame, orient=tk.VERTICAL, command=template_listbox.yview)
            template_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            template_listbox.config(yscrollcommand=template_scrollbar.set)
            
            for header in template_headers:
                template_listbox.insert(tk.END, header)
                
            # Zuordnungen (Rechts)
            mapping_frame_right = ttk.LabelFrame(columns_frame, text="Aktuelle Zuordnungen", padding="10")
            mapping_frame_right.grid(row=0, column=2, sticky=(tk.N, tk.S, tk.W, tk.E))
            
            mapping_listbox = tk.Listbox(mapping_frame_right, height=25, width=50, selectmode=tk.SINGLE)
            mapping_listbox.pack(fill=tk.BOTH, expand=True)
            
            # Scrollbar für Mapping-Liste
            mapping_scrollbar = ttk.Scrollbar(mapping_frame_right, orient=tk.VERTICAL, command=mapping_listbox.yview)
            mapping_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            mapping_listbox.config(yscrollcommand=mapping_scrollbar.set)
            
            # Event-Handler für Spalten-Auswahl
            def on_data_select(event):
                selection = data_listbox.curselection()
                if selection:
                    selected_text = data_listbox.get(selection[0])
                    # Entferne "(keine Daten)" falls vorhanden
                    if " (keine Daten)" in selected_text:
                        self.selected_data_col = selected_text.replace(" (keine Daten)", "")
                        self.current_data_selection.config(text=f"Daten-Spalte: {self.selected_data_col} ⚠", foreground='red')
                    else:
                        self.selected_data_col = selected_text
                        self.current_data_selection.config(text=f"Daten-Spalte: {self.selected_data_col}", foreground='green')
                        
            def on_template_select(event):
                selection = template_listbox.curselection()
                if selection:
                    self.selected_template_col = template_listbox.get(selection[0])
                    self.current_template_selection.config(text=f"Vorlage-Spalte: {self.selected_template_col}", foreground='green')
            
            # Doppelklick-Events
            data_listbox.bind('<Double-1>', on_data_select)
            data_listbox.bind('<Button-1>', on_data_select)  # Auch bei einfachem Klick
            template_listbox.bind('<Double-1>', on_template_select)
            template_listbox.bind('<Button-1>', on_template_select)  # Auch bei einfachem Klick
            
            # Grid-Konfiguration für Spalten
            columns_frame.columnconfigure(0, weight=1)
            columns_frame.columnconfigure(1, weight=1)
            columns_frame.columnconfigure(2, weight=1)
            columns_frame.rowconfigure(0, weight=1)
            
            # Button Frame
            button_frame = ttk.Frame(mapping_frame)
            button_frame.pack(fill=tk.X, pady=(15, 0))
            
            # Linke Buttons
            left_buttons = ttk.Frame(button_frame)
            left_buttons.pack(side=tk.LEFT)
            
            ttk.Button(left_buttons, text="Zuordnung hinzufügen", 
                      command=lambda: self.add_mapping_gui_improved(mapping_listbox)).pack(side=tk.LEFT, padx=(0, 10))
            ttk.Button(left_buttons, text="Zuordnung entfernen", 
                      command=lambda: self.remove_mapping_gui(mapping_listbox)).pack(side=tk.LEFT, padx=(0, 10))
            ttk.Button(left_buttons, text="Alle entfernen", 
                      command=lambda: self.clear_all_mappings_gui(mapping_listbox)).pack(side=tk.LEFT, padx=(0, 10))
            
            # Rechte Buttons
            right_buttons = ttk.Frame(button_frame)
            right_buttons.pack(side=tk.RIGHT)
            
            ttk.Button(right_buttons, text="Automatisches Mapping", 
                      command=lambda: self.auto_mapping_gui(data_headers, template_headers, mapping_listbox)).pack(side=tk.LEFT, padx=(0, 10))
            ttk.Button(right_buttons, text="Abbrechen", 
                      command=mapping_window.destroy).pack(side=tk.LEFT, padx=(0, 10))
            ttk.Button(right_buttons, text="Speichern", 
                      command=lambda: self.save_mapping_gui(mapping_listbox, mapping_window), 
                      style='Accent.TButton').pack(side=tk.LEFT)
            
            # Lade vorhandene Mappings
            self.load_existing_mappings_gui(mapping_listbox, data_headers, template_headers)
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Erstellen des Mapping-Fensters: {str(e)}")
            
    def add_mapping_gui_improved(self, mapping_listbox):
        """Verbesserte Zuordnungs-Funktion mit besserer Auswahl-Behandlung"""
        if not self.selected_data_col or not self.selected_template_col:
            messagebox.showwarning("Warnung", "Bitte wählen Sie sowohl eine Daten- als auch eine Vorlage-Spalte aus.")
            return
        
        # Prüfe ob Daten-Spalte verfügbar ist
        if self.available_data_columns and self.selected_data_col not in self.available_data_columns:
            result = messagebox.askyesno("Warnung", 
                f"Die Spalte '{self.selected_data_col}' enthält keine Daten. Trotzdem zuordnen?")
            if not result:
                return
        
        # Prüfe auf Duplikate
        existing_mappings = [mapping_listbox.get(i) for i in range(mapping_listbox.size())]
        new_mapping = f"{self.selected_data_col} → {self.selected_template_col}"
        
        # Prüfe ob diese Daten-Spalte bereits zugeordnet ist
        if any(mapping.startswith(f"{self.selected_data_col} →") for mapping in existing_mappings):
            messagebox.showwarning("Warnung", f"Die Spalte '{self.selected_data_col}' ist bereits zugeordnet.")
            return
            
        # Prüfe ob diese Template-Spalte bereits zugeordnet ist
        if any(mapping.endswith(f"→ {self.selected_template_col}") for mapping in existing_mappings):
            messagebox.showwarning("Warnung", f"Die Vorlage-Spalte '{self.selected_template_col}' ist bereits zugeordnet.")
            return
            
        mapping_listbox.insert(tk.END, new_mapping)
        
        # Reset Auswahl
        self.selected_data_col = None
        self.selected_template_col = None
        self.current_data_selection.config(text="Daten-Spalte: Keine ausgewählt", foreground='gray')
        self.current_template_selection.config(text="Vorlage-Spalte: Keine ausgewählt", foreground='gray')
        
    def remove_mapping_gui(self, mapping_listbox):
        selection = mapping_listbox.curselection()
        if selection:
            mapping_listbox.delete(selection[0])
        else:
            messagebox.showwarning("Warnung", "Bitte wählen Sie eine Zuordnung zum Entfernen aus.")
            
    def clear_all_mappings_gui(self, mapping_listbox):
        """Löscht alle Zuordnungen nach Bestätigung"""
        if mapping_listbox.size() > 0:
            result = messagebox.askyesno("Bestätigung", "Möchten Sie wirklich alle Zuordnungen löschen?")
            if result:
                mapping_listbox.delete(0, tk.END)
        
    def auto_mapping_gui(self, data_headers, template_headers, mapping_listbox):
        mapping_listbox.delete(0, tk.END)
        
        mapped_templates = set()
        
        # Zuerst exakte Übereinstimmungen
        for data_header in data_headers:
            if data_header in template_headers and data_header not in mapped_templates:
                # Nur hinzufügen wenn Daten verfügbar sind
                if not self.available_data_columns or data_header in self.available_data_columns:
                    mapping_listbox.insert(tk.END, f"{data_header} → {data_header}")
                    mapped_templates.add(data_header)
        
        # Dann Keyword-basiertes Mapping
        for data_header in data_headers:
            # Überspringe wenn bereits gemappt oder keine Daten verfügbar
            existing_mappings = [mapping_listbox.get(i) for i in range(mapping_listbox.size())]
            if any(mapping.startswith(f"{data_header} →") for mapping in existing_mappings):
                continue
                
            if self.available_data_columns and data_header not in self.available_data_columns:
                continue
                
            # Prüfe auf Keyword-Mapping
            for keyword, template_pattern in self.keyword_mappings.items():
                if keyword.lower() in data_header.lower():
                    best_match = self.find_template_match(template_pattern, template_headers)
                    if best_match and best_match not in mapped_templates:
                        mapping_listbox.insert(tk.END, f"{data_header} → {best_match}")
                        mapped_templates.add(best_match)
                        break
            
            # Wenn kein Keyword-Match, versuche ähnliche Spalten
            if not any(mapping.startswith(f"{data_header} →") for mapping in [mapping_listbox.get(i) for i in range(mapping_listbox.size())]):
                best_match = self.find_best_match_gui(data_header, template_headers, mapped_templates)
                if best_match:
                    mapping_listbox.insert(tk.END, f"{data_header} → {best_match}")
                    mapped_templates.add(best_match)
                    
    def find_best_match_gui(self, data_header, template_headers, mapped_templates):
        data_words = set(data_header.lower().split('_'))
        best_score = 0
        best_match = None
        
        for template_header in template_headers:
            if template_header in mapped_templates:
                continue
                
            template_words = set(template_header.lower().split('_'))
            common_words = data_words.intersection(template_words)
            score = len(common_words)
            
            if score > best_score and score > 1:  # Mindestens 2 gemeinsame Wörter
                best_score = score
                best_match = template_header
                
        return best_match
        
    def load_existing_mappings_gui(self, mapping_listbox, data_headers, template_headers):
        # Lade vorhandene Mappings wenn vorhanden
        if self.mapping:
            for data_col, template_col in self.mapping.items():
                # Überprüfe ob die Spalten noch existieren
                if data_col in data_headers and template_col in template_headers:
                    mapping_listbox.insert(tk.END, f"{data_col} → {template_col}")
        else:
            # Führe automatisches Mapping durch
            self.auto_mapping_gui(data_headers, template_headers, mapping_listbox)
        
    def save_mapping_gui(self, mapping_listbox, window):
        self.mapping = {}
        failed_mappings = []
        
        for i in range(mapping_listbox.size()):
            mapping = mapping_listbox.get(i)
            if " → " in mapping:
                try:
                    data_col, template_col = mapping.split(" → ")
                    self.mapping[data_col] = template_col
                except ValueError:
                    failed_mappings.append(mapping)
            
        # Update Status
        count = len(self.mapping)
        if count > 0:
            self.mapping_status.config(
                text=f"✓ {count} Zuordnungen gespeichert", 
                foreground='green'
            )
            success_msg = f"{count} Zuordnungen gespeichert."
            if failed_mappings:
                success_msg += f"\n{len(failed_mappings)} Zuordnungen konnten nicht verarbeitet werden."
            messagebox.showinfo("Erfolg", success_msg)
        else:
            self.mapping_status.config(
                text="⚠ Keine Zuordnungen gespeichert", 
                foreground='orange'
            )
            
        window.destroy()
        
    def read_headers(self, file_path, header_row, sheet_name):
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row-1, nrows=0)
            return list(df.columns)
        except Exception as e:
            raise Exception(f"Fehler beim Lesen der Header von {file_path}: {str(e)}")
            
    def pattern_match_score(self, pattern, text):
        try:
            # Ersetze Platzhalter für Regex
            regex_pattern = pattern.replace(r'\[.*\]', r'\[[^\]]*\]')
            regex_pattern = regex_pattern.replace('*', '.*')
            regex_pattern = regex_pattern.replace('#', r'\#')
            
            # Entferne die Escape-Zeichen für den Punkt
            regex_pattern = regex_pattern.replace(r'\.', '.')
            
            # Führe Regex-Matching durch
            if re.fullmatch(regex_pattern, text):
                # Zähle die Übereinstimmungen von Schlüsselwörtern
                pattern_keywords = set(re.findall(r'[a-zA-Z_]+', pattern))
                text_keywords = set(re.findall(r'[a-zA-Z_]+', text))
                common_keywords = pattern_keywords.intersection(text_keywords)
                return len(common_keywords)
            return 0
        except Exception as e:
            print(f"Regex-Fehler bei Pattern '{pattern}': {e}")
            return 0
        
    def show_debug_info(self):
        if not self.template_file.get() or not self.data_file.get():
            messagebox.showerror("Fehler", "Bitte wählen Sie beide Dateien aus.")
            return
            
        try:
            # Lade Header aus beiden Dateien
            template_headers = self.read_headers(
                self.template_file.get(), 
                self.template_header_row.get(),
                self.template_sheet.get()
            )
            data_headers = self.read_headers(
                self.data_file.get(), 
                self.data_header_row.get(),
                self.data_sheet.get()
            )
            
            # Debug-Info zusammenstellen
            debug_info = f"""DEBUG-INFORMATIONEN
{'='*50}

KONFIGURATION:
- Template: {self.template_file.get()} (Blatt: {self.template_sheet.get()})
- Daten: {self.data_file.get()} (Blatt: {self.data_sheet.get()})
- Template Header-Zeile: {self.template_header_row.get()}, Start-Zeile: {self.template_start_row.get()}
- Daten Header-Zeile: {self.data_header_row.get()}, Start-Zeile: {self.data_start_row.get()}

VERFÜGBARE DATEN:
- Validiert: {'Ja' if self.available_data_columns else 'Nein'}
- Spalten mit Daten: {len(self.available_data_columns)}
{chr(10).join(f"  - {col}" for col in sorted(self.available_data_columns)) if self.available_data_columns else "  (Noch nicht validiert)"}

DATEN-HEADER ({len(data_headers)}):
{chr(10).join(f"  {i+1}. {header}" for i, header in enumerate(data_headers))}

TEMPLATE-HEADER ({len(template_headers)}):
{chr(10).join(f"  {i+1}. {header}" for i, header in enumerate(template_headers))}

AKTUELLE MAPPINGS ({len(self.mapping)}):
{chr(10).join(f"  {data_col} → {template_col}" for data_col, template_col in self.mapping.items()) if self.mapping else "  (Keine Mappings erstellt)"}

KEYWORD-MAPPINGS:
{chr(10).join(f"  {keyword} → {pattern}" for keyword, pattern in list(self.keyword_mappings.items())[:10])}
  ... und {len(self.keyword_mappings)-10} weitere
"""
            
            # Zeige Debug-Info in separatem Fenster
            debug_window = tk.Toplevel(self.root)
            debug_window.title("Debug-Informationen")
            debug_window.geometry("900x700")
            debug_window.grab_set()
            
            # Text-Widget mit Scrollbar
            frame = ttk.Frame(debug_window, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            text_widget = tk.Text(frame, wrap=tk.WORD, font=('Consolas', 9))
            text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            text_widget.config(yscrollcommand=scrollbar.set)
            
            text_widget.insert(tk.END, debug_info)
            text_widget.config(state=tk.DISABLED)
            
            # Schließen-Button
            ttk.Button(debug_window, text="Schließen", command=debug_window.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Erstellen der Debug-Info: {str(e)}")
        
    def start_migration(self):
        if not self.template_file.get() or not self.data_file.get():
            messagebox.showerror("Fehler", "Bitte wählen Sie beide Dateien aus.")
            return
        
        if not self.template_sheet.get() or not self.data_sheet.get():
            messagebox.showerror("Fehler", "Bitte wählen Sie Tabellenblätter aus.")
            return
        
        if not self.mapping:
            messagebox.showerror("Fehler", "Bitte erstellen Sie zuerst ein Mapping.")
            return
        
        try:
            self.progress.start()
            self.status_label.config(text="Migration läuft...", foreground='blue')
            self.migration_button.config(state='disabled')
            self.root.update()
        
            # Lade Daten
            data_df = pd.read_excel(
                self.data_file.get(), 
                sheet_name=self.data_sheet.get(),
                header=self.data_header_row.get()-1
            )
        
            # Überspringe die Zeilen vor der Startzeile
            if self.data_start_row.get() > self.data_header_row.get() + 1:
                skip_rows = self.data_start_row.get() - (self.data_header_row.get() + 1)
                data_df = data_df.iloc[skip_rows:]
        
            # Debug-Ausgabe
            print(f"Anzahl der Zeilen in Daten: {len(data_df)}")
            if len(data_df) > 0:
                print(f"Erste Zeile der Daten: {data_df.iloc[0].to_dict()}")
        
            # Lade Vorlage MIT VBA-Unterstützung
            template_wb = load_workbook(self.template_file.get(), keep_vba=True)
            template_ws = template_wb[self.template_sheet.get()]
        
            # Debug: Zeige alle Spaltenüberschriften in der Vorlage
            print("Vorlage-Spaltenüberschriften:")
            template_headers = {}
            for col in range(1, template_ws.max_column + 1):
                header_value = template_ws.cell(row=self.template_header_row.get(), column=col).value
                if header_value:
                    template_headers[col] = header_value
                    print(f"Spalte {col}: {header_value}")
        
            # Schreibe Daten in Vorlage
            template_start_row = self.template_start_row.get()
            migrated_rows = 0
            migration_log = []
            skipped_mappings = []
        
            for data_idx, data_row in data_df.iterrows():
                template_row_idx = template_start_row + data_idx
                row_migrated = False
            
                for data_col, template_col in self.mapping.items():
                    if data_col in data_df.columns:
                        value = data_row[data_col]
                    
                    # Finde die Spalte in der Vorlage
                    col_idx = None
                    for idx, header_value in template_headers.items():
                        if header_value == template_col:
                            col_idx = idx
                            break
                            
                    if col_idx is not None and pd.notna(value):
                        template_ws.cell(row=template_row_idx, column=col_idx, value=value)
                        row_migrated = True
                        migration_log.append(f"Zeile {template_row_idx}, Spalte {col_idx}: {data_col} → {template_col} = {value}")
                    elif col_idx is None:
                        skipped_mappings.append(f"Template-Spalte '{template_col}' nicht gefunden für Mapping '{data_col}' → '{template_col}'")
            
                if row_migrated:
                    migrated_rows += 1
        
            # Debug: Zeige Migrations-Log
            print(f"\nMigrations-Log (erste 10 Einträge von {len(migration_log)}):")
            for log_entry in migration_log[:10]:
                print(log_entry)
            
            if skipped_mappings:
                print(f"\nÜbersprungene Mappings ({len(skipped_mappings)}):")
                for skip in skipped_mappings[:5]:
                    print(skip)
        
            # Speichere die migrierte Datei
            output_file = filedialog.asksaveasfilename(
                title="Migrierte Datei speichern",
                defaultextension=".xlsm",
                filetypes=[("Excel Macro-Enabled Workbook", "*.xlsm"), 
                          ("Excel Workbook", "*.xlsx"),
                          ("All files", "*.*")]
            )
        
            if output_file:
                # Stelle sicher, dass die Datei die korrekte Endung hat
                if not output_file.lower().endswith('.xlsm'):
                    output_file += '.xlsm'
            
                # Speichere mit expliziter Angabe des Dateityps
                template_wb.save(output_file)
            
                # Erfolgs-/Warnmeldung
                success_msg = f"Migration abgeschlossen!\n\n"
                success_msg += f"• {migrated_rows} Zeilen migriert\n"
                success_msg += f"• {len(migration_log)} Werte übertragen\n"
                if skipped_mappings:
                    success_msg += f"• {len(skipped_mappings)} Mappings übersprungen (siehe Debug-Info)\n"
                success_msg += f"\nDatei gespeichert als:\n{output_file}"
            
                messagebox.showinfo("Migration abgeschlossen", success_msg)
            
                self.status_label.config(text=f"✓ Migration erfolgreich: {migrated_rows} Zeilen verarbeitet", foreground='green')
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler während der Migration: {str(e)}")
            self.status_label.config(text=f"✗ Migration fehlgeschlagen: {str(e)}", foreground='red')
            import traceback
            traceback.print_exc()
        finally:
            self.progress.stop()
            self.migration_button.config(state='normal')

if __name__ == "__main__":
    root = tk.Tk()
    
    # Konfiguriere Style für bessere Optik
    style = ttk.Style()
    style.configure('Accent.TButton', foreground='white')
    
    app = ExcelMigrationTool(root)
    root.mainloop()
